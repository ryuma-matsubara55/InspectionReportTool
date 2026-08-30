from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from datetime import datetime
from PIL import Image
import io
import os

class ExcelGenerator:
    def __init__(self):
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _get_image_size(self, img_data):
        """画像データから(width, height)を取得する"""
        try:
            if isinstance(img_data, bytes):
                with Image.open(io.BytesIO(img_data)) as img:
                    return img.size
            elif isinstance(img_data, str) and os.path.exists(img_data):
                with Image.open(img_data) as img:
                    return img.size
        except Exception as e:
            print(f"[WARN] Failed to get image size: {e}")
        return (0, 0)

    def create_excel(self, test_cases_data, output_path):
        """テストケースデータからExcelファイルを生成"""
        # write_only=True でメモリ使用量を抑制
        wb = Workbook(write_only=True)
        
        # シートごとにデータを振り分け
        sheet_data = {}
        for data in test_cases_data:
            sheet_name = data.get('sheet', 'シート1')
            if sheet_name not in sheet_data:
                sheet_data[sheet_name] = []
            sheet_data[sheet_name].append(data)

        # 各シートを作成
        for sheet_name, cases in sheet_data.items():
            ws = wb.create_sheet(title=sheet_name)
            self.write_sheet(ws, cases)

        wb.save(output_path)

    def write_sheet(self, ws, test_cases):
        """シートにデータを書き込む"""
        max_inputs = 0
        max_results = 0
        
        # 画像サイズ情報の事前計算
        # 列ごとの最大幅(px)と、テストケースごとの最大高さ(px)を記録
        input_max_widths = {}
        result_max_widths = {}
        tc_image_heights = {}

        for i, data in enumerate(test_cases):
            inputs = data.get('input_images', [])
            results = data.get('result_images', [])
            
            max_inputs = max(max_inputs, len(inputs))
            max_results = max(max_results, len(results))
            
            max_h = 0
            for j, img in enumerate(inputs):
                w, h = self._get_image_size(img)
                input_max_widths[j] = max(input_max_widths.get(j, 10), w)
                max_h = max(max_h, h)
                
            for j, img in enumerate(results):
                w, h = self._get_image_size(img)
                result_max_widths[j] = max(result_max_widths.get(j, 10), w)
                max_h = max(max_h, h)
                
            tc_image_heights[i] = max_h

        # ヘッダー構築
        headers = ['No', '項目', '入力条件', '操作手順', '期待結果', '合否', '実施者', '実施日', 'メモ']
        
        input_cols_start = len(headers) + 1 # 1-indexed
        for j in range(max_inputs):
            headers.append(f'入力画像{j+1}')
            
        result_cols_start = len(headers) + 1
        for j in range(max_results):
            headers.append(f'結果画像{j+1}')

        # 列幅設定 (A~Iまで)
        widths = {
            'A': 5, 'B': 30, 'C': 35, 'D': 45,
            'E': 45, 'F': 8, 'G': 10, 'H': 12, 'I': 30
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
            
        # 画像列の列幅設定（画像の最大幅に合わせて拡縮）
        # 1文字分 ≈ 7px として換算。ある程度の余裕(パディング)を追加
        for j in range(max_inputs):
            col_letter = get_column_letter(input_cols_start + j)
            w_chars = (input_max_widths.get(j, 100) / 7.0) + 1.5
            ws.column_dimensions[col_letter].width = w_chars
            
        for j in range(max_results):
            col_letter = get_column_letter(result_cols_start + j)
            w_chars = (result_max_widths.get(j, 100) / 7.0) + 1.5
            ws.column_dimensions[col_letter].width = w_chars

        current_row_idx = 1
        
        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border_thin
            header_cells.append(cell)
        
        ws.append(header_cells)
        current_row_idx += 1

        for i, data in enumerate(test_cases):
            meta = {
                'tc_idx': i,
                'max_inputs': max_inputs,
                'max_results': max_results,
                'input_cols_start': input_cols_start,
                'result_cols_start': result_cols_start,
                'input_max_widths': input_max_widths,
                'result_max_widths': result_max_widths,
                'max_img_h': tc_image_heights.get(i, 0)
            }
            rows_added = self.write_test_case(ws, data, current_row_idx, meta)
            current_row_idx += rows_added

    def write_test_case(self, ws, data, start_row, meta=None):
        """1つのテストケースをシートに書き込む"""
        
        test_case_number = data.get('number', 1)
        if test_case_number % 2 == 0:
            bg_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        else:
            bg_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        expected_results = data.get('expected_results', [])
        if not expected_results:
            expected_results = [{
                'expected': data.get('expected', ''),
                'result': data.get('result', '未実施'),
                'executor': '',
                'date': '',
                'memo': data.get('result_memo', '')
            }]
            
        row_span = len(expected_results)
        
        if 'item' in data:
            item_value = data['item']
        else:
            summary = data.get('summary', '')
            detail = data.get('detail', '')
            if summary and detail:
                item_value = f"{summary}\n\n{detail}"
            elif summary:
                item_value = summary
            elif detail:
                item_value = detail
            else:
                item_value = ''
        
        common_values = [
            data['number'],
            item_value,
            data['input_condition'],
            data['procedure']
        ]
        
        # 結合処理 (A〜D列のみ。画像の結合は廃止し1セルに1画像とする)
        if row_span > 1:
            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                ws.merged_cells.add(f'{col_letter}{start_row}:{col_letter}{start_row + row_span - 1}')

        common_height_needed = 0
        if row_span > 0:
            h_item = self._calculate_row_height(common_values[1], 30)
            h_input = self._calculate_row_height(common_values[2], 35)
            h_proc = self._calculate_row_height(common_values[3], 45)
            common_height_needed = max(h_item, h_input, h_proc)

        row_heights = []
        for res in expected_results:
            h_exp = self._calculate_row_height(res.get('expected', ''), 45)
            h_memo = self._calculate_row_height(res.get('memo', ''), 30)
            row_heights.append(max(h_exp, h_memo))
        
        # 期待結果の1行目(start_row)の高さに、テストケース内での最大画像高さを反映させる
        if meta and meta.get('max_img_h', 0) > 0:
            # 高さはポイント換算 (1px ≈ 0.75pt). 余白のため少し足す
            img_h_pt = meta['max_img_h'] * 0.75 + 10
            row_heights[0] = max(row_heights[0], img_h_pt)
            
        total_height = sum(row_heights)
        if total_height < common_height_needed:
            diff = common_height_needed - total_height
            add_per_row = diff / len(row_heights)
            row_heights = [h + add_per_row for h in row_heights]

        for i, res in enumerate(expected_results):
            current_r = start_row + i
            row_cells = []
            
            for j, val in enumerate(common_values):
                cell_value = val if i == 0 else ''
                cell = WriteOnlyCell(ws, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.border_thin
                cell.fill = bg_fill
                row_cells.append(cell)

            cell = WriteOnlyCell(ws, value=res.get('expected', ''))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = self.border_thin
            cell.fill = bg_fill
            row_cells.append(cell)
            
            result = res.get('result', '未実施')
            cell = WriteOnlyCell(ws, value=result)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border_thin
            if result == 'OK':
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif result == 'NG':
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            else:
                cell.fill = bg_fill
            row_cells.append(cell)

            cell = WriteOnlyCell(ws, value=res.get('executor', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border_thin
            cell.fill = bg_fill
            row_cells.append(cell)

            cell = WriteOnlyCell(ws, value=res.get('date', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border_thin
            cell.fill = bg_fill
            row_cells.append(cell)
                
            cell = WriteOnlyCell(ws, value=res.get('memo', ''))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = self.border_thin
            row_cells.append(cell)
            
            # 画像用の空セルを作成（1行目のみ画像が後からアンカー配置され、他の行は空となる）
            if meta:
                for j in range(meta['max_inputs']):
                    cell = WriteOnlyCell(ws, value='')
                    cell.border = self.border_thin
                    row_cells.append(cell)
                for j in range(meta['max_results']):
                    cell = WriteOnlyCell(ws, value='')
                    cell.border = self.border_thin
                    row_cells.append(cell)

            ws.row_dimensions[current_r].height = row_heights[i]
            ws.append(row_cells)

        # 画像の配置 (1行目の各画像セルに対して行う)
        if meta:
            input_images = data.get('input_images', [])
            for j, img in enumerate(input_images):
                w_px = meta['input_max_widths'].get(j, 100)
                h_px = meta['max_img_h']
                col_idx = meta['input_cols_start'] + j
                self.write_single_image(ws, img, start_row, col_idx, w_px, h_px)

            result_images = data.get('result_images', [])
            for j, img in enumerate(result_images):
                w_px = meta['result_max_widths'].get(j, 100)
                h_px = meta['max_img_h']
                col_idx = meta['result_cols_start'] + j
                self.write_single_image(ws, img, start_row, col_idx, w_px, h_px)

        return row_span

    def write_single_image(self, ws, img_data, row, col, max_w_px, max_h_px):
        """指定されたセルに縦横比を維持して画像を貼り付ける(contain)"""
        try:
            import io, os
            
            img_bytes = None
            if isinstance(img_data, bytes):
                img_bytes = io.BytesIO(img_data)
            elif isinstance(img_data, str):
                if not os.path.exists(img_data):
                    return
                img_bytes = img_data
            else:
                return
                
            xl_img = XLImage(img_bytes)
            
            orig_w = xl_img.width
            orig_h = xl_img.height
            if orig_w > 0 and orig_h > 0 and max_w_px > 0 and max_h_px > 0:
                scale_w = max_w_px / orig_w
                scale_h = max_h_px / orig_h
                scale = min(scale_w, scale_h)
                
                # 枠線に被らないように少しパディング
                target_w = int(orig_w * scale) - 4
                target_h = int(orig_h * scale) - 4
                
                xl_img.width = max(1, target_w)
                xl_img.height = max(1, target_h)
            
            # アンカー設定 (セル内に配置)
            col_letter = get_column_letter(col)
            
            # AnchorMarker を使わず、シンプルなセル参照アンカーで十分配置可能
            # 若干のオフセットが必要であれば TwoCellAnchor などを使うが、
            # 左上端合わせならこれで問題ない
            xl_img.anchor = f'{col_letter}{row}'
            
            ws.add_image(xl_img)
            
        except Exception as e:
            print(f"[ERROR] Error loading image: {e}")
            import traceback
            traceback.print_exc()

    def _calculate_row_height(self, text, char_per_line=80):
        """テキストの長さから行の高さを計算"""
        if not text:
            return 20
        lines = text.count('\n') + 1
        for line in text.split('\n'):
            length = 0
            for char in line:
                if ord(char) > 255:
                    length += 2
                else:
                    length += 1
            lines += (length - 1) // char_per_line
            
        return max(20, lines * 15)

    def _apply_border_to_merged_cells(self, ws, cell_range):
        """結合セルに枠線を適用する"""
        pass

    def load_excel(self, file_path, progress_callback=None):
        """Excelファイルからデータを読み込む"""
        if progress_callback:
            progress_callback(0, 100, "Excelファイルを解析中...")
            
        wb = load_workbook(file_path, read_only=False, data_only=True)
        all_data = []

        try:
            total_sheets = len(wb.sheetnames)
            for i, sheet_name in enumerate(wb.sheetnames):
                if progress_callback:
                    # ワークブックの読み込みに時間がかかるため、シート処理は10%~90%の間に割り当てる
                    progress = 10 + int((i / max(1, total_sheets)) * 80)
                    progress_callback(progress, 100, f"シート '{sheet_name}' を読み込み中...")
                    
                ws = wb[sheet_name]
                sheet_cases = self.read_sheet(ws, sheet_name)
                all_data.extend(sheet_cases)
                
            if progress_callback:
                progress_callback(90, 100, "データの構築中...")
        finally:
            wb.close()

        return all_data

    def read_sheet(self, ws, sheet_name):
        """シートからテストケースデータを読み込む（動的列レイアウト対応）"""
        test_cases = []
        
        images_map = {}
        images = getattr(ws, '_images', [])
        
        for img in images:
            try:
                anchor = img.anchor
                marker = getattr(anchor, '_from', anchor)
                # バージョンによってマーカーが取れない場合は別の取得方法を試みる
                if hasattr(marker, 'col') and hasattr(marker, 'row'):
                    col = getattr(marker, 'col', 0) + 1
                    row = getattr(marker, 'row', 0) + 1
                else:
                    # 'K5' のようなセル文字列からのパース (簡易的)
                    try:
                        import re
                        m = re.match(r'([a-zA-Z]+)(\d+)', str(anchor))
                        if m:
                           from openpyxl.utils.cell import column_index_from_string
                           col = column_index_from_string(m.group(1))
                           row = int(m.group(2))
                        else:
                            continue
                    except:
                        continue
                
                img_bytes = None
                if hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                    img.ref.seek(0)
                    img_bytes = img.ref.read()
                elif hasattr(img, '_data'):
                     img_bytes = img._data()
                else:
                    img_pil = img.ref
                    if img_pil:
                        buf = io.BytesIO()
                        img_pil.save(buf, format='PNG')
                        img_bytes = buf.getvalue()
                
                if img_bytes:
                    key = (row, col)
                    if key not in images_map:
                        images_map[key] = []
                    images_map[key].append(img_bytes)
            except Exception as e:
                print(f"[WARN] Failed to extract image: {e}")
                continue

        rows = ws.iter_rows(values_only=True)
        
        header_row_idx = None
        current_row_idx = 0
        input_image_cols = []
        result_image_cols = []
        
        for row in rows:
            current_row_idx += 1
            if row[0] == 'No':
                header_row_idx = current_row_idx
                for col_idx, val in enumerate(row):
                    if isinstance(val, str):
                        if val.startswith('入力画像'):
                            input_image_cols.append(col_idx + 1)
                        elif val.startswith('結果画像'):
                            result_image_cols.append(col_idx + 1)
                break
        
        if not header_row_idx:
            return test_cases
        
        current_case = None
        for row in rows:
            current_row_idx += 1
            no_value = row[0]
            
            if no_value:
                if current_case and current_case.get('expected_results'):
                    test_cases.append(current_case)
                    current_case = None
                
                try:
                    number = int(no_value)
                except:
                    number = len(test_cases) + 1
                
                current_case = {
                    'number': number,
                    'item': row[1] or '',
                    'input_condition': row[2] or '',
                    'procedure': row[3] or '',
                    'expected_results': [],
                    'sheet': sheet_name,
                    'input_images': [],
                    'result_images': []
                }
                
                for c in input_image_cols:
                    if (current_row_idx, c) in images_map:
                        current_case['input_images'].extend(images_map[(current_row_idx, c)])
                for c in result_image_cols:
                    if (current_row_idx, c) in images_map:
                        current_case['result_images'].extend(images_map[(current_row_idx, c)])
            
            if current_case:
                expected = row[4] or ''
                result = row[5] or '未実施'
                executor = row[6] or ''
                date = row[7] or ''
                memo = row[8] or ''
                
                if expected or result != '未実施':
                    current_case['expected_results'].append({
                        'expected': expected,
                        'result': result,
                        'executor': executor,
                        'date': date,
                        'memo': memo
                    })
                
                # 結合等で次行にアンカー判定がずれた場合なども救済
                if not no_value:
                    for c in input_image_cols:
                        if (current_row_idx, c) in images_map:
                            current_case['input_images'].extend(images_map[(current_row_idx, c)])
                    for c in result_image_cols:
                        if (current_row_idx, c) in images_map:
                            current_case['result_images'].extend(images_map[(current_row_idx, c)])

        if current_case and current_case.get('expected_results'):
            test_cases.append(current_case)
        
        return test_cases
